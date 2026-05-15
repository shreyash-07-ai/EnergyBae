import streamlit as st
import google.generativeai as genai
from PIL import Image
import pandas as pd
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Configure Gemini API
GEMINI_API_KEY = "AIzaSyC5NQNdW9dAC209d8SUj3HVNLqNo2CNAOw"                                                    # Gemini API Key By mail shreyashmusmade123@gmail.com   
genai.configure(api_key=GEMINI_API_KEY)

st.set_page_config(page_title="Electricity Bill to Excel", layout="wide")
st.title("Electricity Bill to Excel")                                                                         # Title of the web

uploaded_file = st.file_uploader("Upload Electricity Bill Image", type=["jpg", "jpeg", "png"])                # File uploader button

def extract_json(text):
    match = re.search(r"\{.*\}", text, re.DOTALL)                                                             # extract JSON from the text using regex. 
    if match:                                                                                                 # re.dotall allows new data
        return match.group(0)
    return None

if uploaded_file:
    image = Image.open(uploaded_file)                                                                         # open uploaded image using PIL
    st.image(image, caption="Uploaded Bill", use_container_width=True)

    if st.button("Extract Data and Generate Excel"):                                                         # buttlon to extract the data
        model = genai.GenerativeModel("gemini-2.5-flash")                                                    # Gemini model used to extract data

        prompt = """                                                                                        
        Read the electricity bill image and extract the following information in JSON format only.

        {
          "consumer_name": "",
          "consumer_number": "",
          "phone_or_reference": "",
          "bill_type": "",
          "connection_type": "",
          "fixed_charge": "",
          "months": [
            {
              "month": "",
              "units": "",
              "bill_amount": "",
              "unit_cost": ""
            }
          ],
          "average_units": "",
          "average_bill_amount": "",
          "average_unit_cost": "",
          "solar_panels": "",
          "solar_capacity": ""
        }

        Important:
        - Extract connection type (1 phase, 3 phase, Residential, etc.)
        - Extract all monthly usage history from the bill table. 
          Return months in "YYYY-MM" format (e.g., 2026-01) for reliable sorting.
        - Extract average values if shown.
        - Extract solar panel and capacity information if available.
        - If any value is not available, keep it blank.
        - Return only JSON.
        """                                                             # prompt work as an agent

        response = model.generate_content([prompt, image])              # gemini model receive both image and prompt and make response as per prompt
        raw_text = response.text                                            # extracted data from model response 

        json_text = extract_json(raw_text)                                              # extract JSON from the Gemini response using the defined function

        if json_text:
            data = json.loads(json_text)                                        #it converts json text or string in to python dictionary

            # Set fixed charge based on connection type
            connection_type = data.get("connection_type", "").lower()
            if "1" in connection_type and "phase" in connection_type:
                data["fixed_charge"] = "140"
            elif "3" in connection_type and "phase" in connection_type:
                data["fixed_charge"] = "440"                                            # fixed charges as per requirenments

            st.subheader("Extracted Data")
            st.json(data)

            wb = Workbook()                                                         
            ws = wb.active                                                      
            ws.title = "Electricity Bill"                                               # creating excel workbook with title Electricity Bill

            # Define colors
            header_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")  # Orange
            average_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
            peach_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # Peach/Orange
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            bold_font = Font(bold=True, color="FFFFFF")
            red_font = Font(color="FFFFFF", bold=True)
            green_font = Font(color="FFFFFF", bold=True)
                                                                                                                            # giving the colors 
            # Top Details (Header section)
            current_row = 1
            ws[f"A{current_row}"] = "Consumer Name"                                                             # label of column A1 and similarly all are also lables of all column
            ws[f"B{current_row}"] = data.get("consumer_name", "")

            current_row += 1
            ws[f"A{current_row}"] = "Consumer Number"
            ws[f"B{current_row}"] = data.get("consumer_number", "")

            current_row += 1
            ws[f"A{current_row}"] = "Phone/Reference"
            ws[f"B{current_row}"] = data.get("phone_or_reference", "")

            current_row += 1
            ws[f"A{current_row}"] = "Bill Type"
            ws[f"B{current_row}"] = data.get("bill_type", "")

            current_row += 1
            ws[f"A{current_row}"] = "Connection Type"
            ws[f"B{current_row}"] = data.get("connection_type", "")

            current_row += 1
            ws[f"A{current_row}"] = "Fixed Charge"
            fixed_charge = data.get("fixed_charge", "")
            try:
                if fixed_charge not in ["", None]:
                    fixed_charge = float(str(fixed_charge).replace(",", ""))
            except (ValueError, TypeError):
                pass
            ws[f"B{current_row}"] = fixed_charge

            current_row += 1
            ws[f"A{current_row}"] = "Solar Panel Capacity"
            ws[f"B{current_row}"] = "600"                                       # default solar panel capacity as per requirements

            # Add blank row
            current_row += 2

            # Monthly Table Header
            start_row = current_row
            headers = ["Month", "Units", "Bill Amount", "Unit Cost"]                    # headers of the table as per requirements

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_num)
                cell.value = header
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = None

            months = data.get("months", [])                                         

            # Filter and sort months (Rolling 12 months)
            parsed_months = []
            for m in months:
                month_str = m.get("month", "")
                try:
                    # Try parsing YYYY-MM first as requested in prompt
                    dt = datetime.strptime(month_str, "%Y-%m")                                      # date time convert to one single format
                except:
                    try:
                        # Fallback to other common formats if Gemini ignores instructions
                        dt = datetime.strptime(month_str, "%B %Y")
                    except:
                        try:
                            dt = datetime.strptime(month_str, "%b %Y")
                        except:
                            dt = None
                
                if dt:
                    m["_dt"] = dt
                    parsed_months.append(m)

            # Sort by date
            parsed_months.sort(key=lambda x: x["_dt"])                                          #sorting months in ascending order based on date
            
            # Keep only the last 12 months
            if len(parsed_months) > 12:
                parsed_months = parsed_months[-12:]                                             # only 12 months kept 
            
            months = parsed_months

            total_months = len(months)
            for idx, month_data in enumerate(months, start=1):
                row = start_row + idx
                
                # Get values and try to convert to numeric for correct Excel behavior (Averages)
                units = month_data.get("units", "")
                bill_amount = month_data.get("bill_amount", "")
                
                try:
                    if units not in ["", None]:
                        units = float(str(units).replace(",", "").strip())
                    else:
                        units = None
                except (ValueError, TypeError):
                    pass
                
                try:
                    if bill_amount not in ["", None]:
                        bill_amount = float(str(bill_amount).replace(",", "").strip())
                    else:
                        bill_amount = None
                except (ValueError, TypeError):
                    pass

                display_month = month_data.get("_dt").strftime("%B %Y") if "_dt" in month_data else month_data.get("month", "")
                ws.cell(row=row, column=1).value = display_month
                ws.cell(row=row, column=2).value = units
                ws.cell(row=row, column=3).value = bill_amount
                
                # Formula for Unit Cost: (Bill Amount - Fixed Charge) / Units
                # Apply only to the last row as per user request
                # Add check for B{row} > 0 to avoid DIV/0 errors
                if idx == total_months:
                    ws.cell(row=row, column=4).value = f"=IF(AND(ISNUMBER(B{row}), B{row}>0), (C{row}-$B$6)/B{row}, \"\")"

            # Average Row with Excel Formulas
            avg_row = start_row + len(months) + 1
            first_data_row = start_row + 1
            last_data_row = start_row + len(months)
            
            ws.cell(row=avg_row, column=1).value = "Average"
            ws.cell(row=avg_row, column=2).value = f"=AVERAGE(B{first_data_row}:B{last_data_row})"
            ws.cell(row=avg_row, column=3).value = f"=AVERAGE(C{first_data_row}:C{last_data_row})"
            ws.cell(row=avg_row, column=4).value = f"=AVERAGE(D{first_data_row}:D{last_data_row})"

            for col in range(1, 5):
                ws.cell(row=avg_row, column=col).fill = average_fill
                ws.cell(row=avg_row, column=col).font = Font(bold=True)

            # Solar Calculations Section (with elegant gap)
            kw_row = avg_row + 3
            sp_row = kw_row + 1
            sc_row = sp_row + 1
            np_row = sc_row + 1

            # kW Row
            ws.cell(row=kw_row, column=1).value = "kW"
            ws.cell(row=kw_row, column=2).value = f"=(B{avg_row}*12*1.1)/1400"
            
            # Solar Panels Row
            ws.cell(row=sp_row, column=1).value = "Solar Panels"
            ws.cell(row=sp_row, column=2).value = f"=(B{kw_row})/$B$7*1000"

            # Solar capacity Row
            ws.cell(row=sc_row, column=1).value = "Solar capacity"
            ws.cell(row=sc_row, column=1).fill = peach_fill
            ws.cell(row=sc_row, column=2).value = f"=ROUND(B{sp_row},0)*$B$7/1000"
            ws.cell(row=sc_row, column=2).fill = yellow_fill

            # Number of Panels Row
            ws.cell(row=np_row, column=1).value = "Number of Panels"
            ws.cell(row=np_row, column=1).fill = peach_fill
            ws.cell(row=np_row, column=2).value = f"=(B{sc_row})/$B$7*1000"
            ws.cell(row=np_row, column=2).fill = green_fill

            # Auto width
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20

            # Generate filename from consumer name with duplicate handling
            consumer_name = data.get("consumer_name", "electricity_bill").replace(" ", "_")
            consumer_number = data.get("consumer_number", "")
            
            if consumer_number:
                base_filename = f"{consumer_name}_{consumer_number}"
            else:
                base_filename = consumer_name
            
            output_file = f"{base_filename}.xlsx"
            counter = 1
            
            # If file exists, add counter
            while os.path.exists(output_file):
                output_file = f"{base_filename}-{counter}.xlsx" 
                counter += 1
            
            wb.save(output_file)

            st.success("Excel file generated successfully!")

            with open(output_file, "rb") as f:
                st.download_button(
                    label="Download Excel File",
                    data=f,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error("Could not extract JSON from Gemini response.")
